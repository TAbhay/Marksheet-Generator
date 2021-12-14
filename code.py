import os
import csv
import openpyxl
from openpyxl.styles import PatternFill  # Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill  # Connect styles for text
from openpyxl.styles import colors  # Connect colors for text and cells
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
if not os.path.isdir("output"):
     os.mkdir("output")   
positive = 5
negative = -1
result = {} 
def generate_result():
    j =0
    with open("sample_input/responses.csv",'r',newline='') as File:         #the csv file is stored in a File object
            reader=csv.reader(File)                              #csv.reader is used to read a file
            for row in reader:
                key = row[6]
                isAns = 0
                if key=="ANSWER":
                    isAns = 1
                    result["ANSWER"] = {}
                    correctAns = []
                    totalQue = 0
                    for i in range(7,len(row)):
                        correctAns.append([row[i],row[i]])
                        totalQue = totalQue + 1
                    result["ANSWER"]["response"] = correctAns
                    result["ANSWER"]["details"]  = [row[0],row[1],row[2],row[3],row[4],row[5],str(totalQue*positive)+"/"+str(totalQue*positive),row[6],]
                    result["ANSWER"]["stats"]  = [["No.",totalQue,0,0,totalQue],["Marking",positive,negative,0,""],["Total",totalQue*positive,0,"",str(totalQue*positive)+"/"+str(totalQue*positive)]]
                    break
            if isAns==0:
                return "No answer exists"
            for row in reader:
                if j==0:
                    j = j+1
                    continue
                key = row[6]
                if key=="ANSWER":
                    continue
                result[key] = {}
                correctAns  = result["ANSWER"]["response"]
                rightAns = 0
                wrongAns = 0
                notAttempt = 0
                marks = 0
                response = []
                totalQ = 0
                for i in range(7,len(row)):
                    totalQ = totalQ + 1
                    studentAns = row[i]
                    corAns = correctAns[i-7][0]
                    response.append([studentAns,corAns])
                    if len(studentAns.strip())==0:
                        notAttempt = notAttempt + 1
                    elif studentAns==corAns:
                        marks = marks + positive
                        rightAns = rightAns + 1
                    else:
                        marks = marks + negative
                        wrongAns = wrongAns + 1
                result[key]["details"]  = [row[0],row[1],row[2],row[3],row[4],row[5],str(marks)+"/"+str(positive*totalQ),row[6]]
                result[key]["response"] = response
                result[key]["stats"] = [["No.",rightAns,wrongAns,notAttempt,totalQ],["Marking",positive,negative,0,""],["Total",positive*rightAns,negative*wrongAns,"",str(marks)+"/"+str(positive*totalQ)]]

                
def concise_marksheet():
    Header = ['Timestamp','Email address','Google_Score','Name','IITP webmail','Phone (10 digit only)','Score_After_Negative','Roll Number']    
    filePath = "output/concise_marksheet.xlsx"
    wb = openpyxl.Workbook()                            #creating a workbook using openpyxl
    sheet = wb.active 
    col = 1
    for item in Header:
         sheet.cell(row = 1,column = col).value = item
         col = col + 1 
    rw = 2                                
    for key , value in result.items():
        detail = value["details"]
        responses = value["response"]
        col = 1
        for item in detail:
            sheet.cell(row = rw,column = col).value = item
            col = col + 1
        for item in responses:
            sheet.cell(row=rw,column=col).value = item[0]
            col = col + 1
        rw = rw + 1


    wb.save(filePath)


      
def individual_marksheet():
    print("hello")
    for key ,value in result.items():
        filePath = "output/"+key+".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        img = openpyxl.drawing.image.Image('IITP LOGO.png')
        img.anchor = 'A1'
        img.width = 655
        img.height = 82
        ws.add_image(img)
        letter= ['A','B','C','D','E']
        studentName = value["details"][3]
        styleLight = Font(name="Century",size=12, color='000000')
        styleBold = Font(name="Century",size=12, color='000000', bold=True)
        styleBlue = Font(name="Century",size=12, color='0000FF')
        styleRight = Font(name="Century",size=12, color='339933')
        styleWrong = Font(name="Century",size=12, color='FF0000')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for x in letter:
            ws.column_dimensions[x].width=18
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
        ws.cell(row=5,column=1,value='Mark Sheet').alignment = Alignment(horizontal='center')
        ws.cell(row=5,column=1,value='Mark Sheet').font =Font(name="Century",size=18, underline='single', color='000000',bold=True)

        ws.cell(row=6,column=1,value='Name:').font =styleLight
        ws.cell(row=6,column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=6,column=2,value=studentName).font =styleBold

        ws.cell(row=6,column=4,value='Exam:').font =styleLight
        ws.cell(row=6,column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=6,column=5,value='quiz').font =styleBold

        ws.cell(row=7,column=1,value='Roll Number').font =styleLight
        ws.cell(row=7,column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=7,column=2,value=key).font =styleBold
        status = value["stats"]
        subHeader = ["","Right","Wrong","NotAttempt","Max"]
        col = 1
        for item in subHeader:
            ws.cell(row=9,column=col,value=item).font = styleBold
            ws.cell(row=9,column=col,value=item).border = thin_border
            ws.cell(row=9,column=col).alignment = Alignment(horizontal='center')
            col = col + 1

        rw = 10
        for item in status:
            fontscheme = ""
            col = 1
            
            for val in item:
                if col==1:
                    fontscheme = styleBold
                elif col == 2:
                    fontscheme = styleRight
                elif col == 3:
                    fontscheme = styleWrong
                else:
                    fontscheme = styleLight
                ws.cell(row=rw,column=col,value=val).font = fontscheme
                ws.cell(row=rw,column=col,value=val).border = thin_border
                ws.cell(row=rw,column=col).alignment = Alignment(horizontal='center')
                col = col+1
            rw = rw + 1
        ws.cell(row=12,column=5).font = styleBlue
        ansHeader = ["Student Ans","Correct Ans"]

        rw = 15
        
        ws.cell(row=rw,column=1,value="Student Ans").font=styleBold
        ws.cell(row=rw,column=2,value="Correct Ans").font=styleBold
        ws.cell(row=rw,column=4,value="Student Ans").font=styleBold
        ws.cell(row=rw,column=5,value="Correct Ans").font=styleBold
        ws.cell(row=rw,column=1,value="Student Ans").border = thin_border
        ws.cell(row=rw,column=2,value="Correct Ans").border = thin_border
        ws.cell(row=rw,column=4,value="Student Ans").border = thin_border
        ws.cell(row=rw,column=5,value="Correct Ans").border = thin_border
        ws.cell(row=rw,column=1,value="Student Ans").alignment = Alignment(horizontal='center')
        ws.cell(row=rw,column=2,value="Correct Ans").alignment = Alignment(horizontal='center')
        ws.cell(row=rw,column=4,value="Student Ans").alignment = Alignment(horizontal='center')
        ws.cell(row=rw,column=5,value="Correct Ans").alignment = Alignment(horizontal='center')
        
        response = value["response"]
        rw = 16
        for i in range(0,len(response)-3):
            studAns = response[i][0]
            corrAns = response[i][1]
            colorScheme = ""
            if studAns==corrAns:
                colorScheme = styleRight
            else:
                colorScheme = styleWrong
            ws.cell(row=rw,column=1,value=studAns).font = colorScheme
            ws.cell(row=rw,column=2,value=corrAns).font = styleBlue
            ws.cell(row=rw,column=1,value=studAns).alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=2,value=corrAns).alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=1,value=studAns).border = thin_border
            ws.cell(row=rw,column=2,value=corrAns).border = thin_border
            rw = rw + 1
        rw = 16
        for i in range(len(response)-3,len(response)):
            studAns = response[i][0]
            corrAns = response[i][1]
            colorScheme = ""
            if studAns==corrAns:
                colorScheme = styleRight
            else:
                colorScheme = styleWrong
            ws.cell(row=rw,column=4,value=studAns).font = colorScheme
            ws.cell(row=rw,column=5,value=corrAns).font = styleBlue
            ws.cell(row=rw,column=4,value=studAns).alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=5,value=corrAns).alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=4,value=studAns).border = thin_border
            ws.cell(row=rw,column=5,value=corrAns).border = thin_border
            rw = rw + 1

        wb.save(filePath)
                       
    
# sheet.cell.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=True)
#thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
generate_result()

concise_marksheet()
print(result["1401CB01"])
individual_marksheet()