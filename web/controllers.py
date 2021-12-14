import os
import csv
from typing import ParamSpec
import openpyxl
from openpyxl.cell.read_only import EmptyCell
from openpyxl.styles import PatternFill  # Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill  # Connect styles for text
from openpyxl.styles import colors  # Connect colors for text and cells
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import smtplib
from email.message import EmailMessage
import glob
import re
  
result = {}
attend = {}

def makeOutputDir():
    if not os.path.isdir("output"):
     os.mkdir("output")
     
positive = 5
negative = -1
def totalStudent():
    try:
        j = 0
        with open("media/input/master_roll.csv",'r',newline='') as File:        
            reader=csv.reader(File)                             
            for row in reader:
                if j == 0:
                    j = j + 1
                    continue
                attend[row[0]] = [row[1],0]
    except Exception as e:
        return "error occurred : " + e
def generate_result(p,n):
    try:
        negative = -1*abs(float(n))
        positive = float(p)
        print(positive,negative)
        j =0
        totalStudent()
        with open("media/input/responses.csv",'r',newline='') as File:         #the csv file is stored in a File object
                reader=csv.reader(File)                              #csv.reader is used to read a file
                for row in reader:
                    key = row[6]
                    isAns = 0
                    if key=="ANSWER":
                        isAns = 1
                        result["ANSWER"] = {}
                        correctAns = []
                        totalQue = 0
                        for i in range(7,len(row)):
                            correctAns.append([row[i],row[i]])
                            totalQue = totalQue + 1
                        attend["ANSWER"][1] = 1
                        result["ANSWER"]["response"] = correctAns
                        result["ANSWER"]["details"]  = [row[0],row[1],str(totalQue*positive)+"/"+str(totalQue*positive),row[3],row[4],row[5],str(totalQue*positive)+"/"+str(totalQue*positive),row[6],]
                        result["ANSWER"]["stats"]  = [["No.",totalQue,0,0,totalQue],["Marking",positive,negative,0,""],["Total",totalQue*positive,0,"",str(totalQue*positive)+"/"+str(totalQue*positive)]]
                        break
                if isAns==0:
                    return "No answer exists"
                for row in reader:
                    if j==0:
                        j = j+1
                        continue
                    key = row[6]
                    if key=="ANSWER":
                        continue
                    result[key] = {}
                    correctAns  = result["ANSWER"]["response"]
                    rightAns = 0
                    wrongAns = 0
                    notAttempt = 0
                    marks = 0
                    response = []
                    totalQ = 0
                    for i in range(7,len(row)):
                        totalQ = totalQ + 1
                        studentAns = row[i]
                        corAns = correctAns[i-7][0]
                        response.append([studentAns,corAns])
                        if len(studentAns.strip())==0:
                            notAttempt = notAttempt + 1
                        elif studentAns==corAns:
                            marks = marks + positive
                            rightAns = rightAns + 1
                        else:
                            marks = marks + negative
                            wrongAns = wrongAns + 1
                    attend[key][1] = 1
                    result[key]["details"]  = [row[0],row[1],str(positive*rightAns)+"/"+str(positive*totalQ),row[3],row[4],row[5],str(marks)+"/"+str(positive*totalQ),row[6]]
                    result[key]["response"] = response
                    result[key]["stats"] = [["No.",rightAns,wrongAns,notAttempt,totalQ],["Marking",positive,negative,0,""],["Total",positive*rightAns,negative*wrongAns,"",str(marks)+"/"+str(positive*totalQ)]]
        # for roll,value in attend:
        #     if value[1]==0:
        #             result[roll]["details"]  = ['','','',value[0],'','','',roll]
        #             result[roll]["response"] = [['',''],['',''],['',''],['',''],['','']]
        #             result[roll]["stats"] = [["No.",'','','',''],["Marking",'','',0,""],["Total",'','',"","ABSENT"]]
        return result
    except Exception as e:
        return "error occurred : " + e

                
def concise_marksheet():
    try:
        totalStudent()
        Header = ['Timestamp','Email address','Google_Score','Name','IITP webmail','Phone (10 digit only)','Score_After_Negative','Roll Number']    
        filePath = "output/concise_marksheet.xlsx"
        wb = openpyxl.Workbook()                            #creating a workbook using openpyxl
        sheet = wb.active 
        col = 1
        for item in Header:
            sheet.cell(row = 1,column = col).value = item
            col = col + 1 
        rw = 2                                
        for key , keyValue in attend.items():
            if not key in result:
                name = keyValue[0]
                roll = key
                sheet.cell(row = rw,column = 1).value = roll
                sheet.cell(row = rw,column = 2).value = name
                sheet.cell(row = rw,column = 3).value = "ABSENT"
                rw = rw+1
                continue
            value = result[key]
            detail = value["details"]
            responses = value["response"]
            col = 1
            for item in detail:
                sheet.cell(row = rw,column = col).value = item
                col = col + 1
            for item in responses:
                sheet.cell(row=rw,column=col).value = item[0]
                col = col + 1
            rw = rw + 1


        wb.save(filePath)
    except Exception as e:
        return "error occurred : "+e


      
def individual_marksheet():
    try:
        totalStudent()
        for key ,keyValue in attend.items():
            filePath = "output/"+key+".xlsx"
            wb = openpyxl.Workbook()
            ws = wb.worksheets[0]
            img = openpyxl.drawing.image.Image('IITP LOGO.png')
            img.anchor = 'A1'
            img.width = 655
            img.height = 82
            ws.add_image(img)
            letter= ['A','B','C','D','E']
            if not key in result:
                studentName = keyValue[0]
            else:
                studentName = result[key]["details"][3] 
            styleLight = Font(name="Century",size=12, color='000000')
            styleBold = Font(name="Century",size=12, color='000000', bold=True)
            styleBlue = Font(name="Century",size=12, color='0000FF')
            styleRight = Font(name="Century",size=12, color='339933')
            styleWrong = Font(name="Century",size=12, color='FF0000')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for x in letter:
                ws.column_dimensions[x].width=18
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
            ws.cell(row=5,column=1,value='Mark Sheet').alignment = Alignment(horizontal='center')
            ws.cell(row=5,column=1,value='Mark Sheet').font =Font(name="Century",size=18, underline='single', color='000000',bold=True)
            
            ws.cell(row=6,column=1,value='Name:').font =styleLight
            ws.cell(row=6,column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=6,column=2,value=studentName).font =styleBold

            ws.cell(row=6,column=4,value='Exam:').font =styleLight
            ws.cell(row=6,column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=6,column=5,value='quiz').font =styleBold

            ws.cell(row=7,column=1,value='Roll Number').font =styleLight
            ws.cell(row=7,column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=7,column=2,value=key).font =styleBold
            if not key in result:
                ws.cell(row=9,column=5,value="ABSENT").font = styleBold
                ws.cell(row=9,column=1,value="ABSENT").font = styleBold
                wb.save(filePath)
                continue
            value = result[key]
            status = value["stats"]
            subHeader = ["","Right","Wrong","NotAttempt","Max"]
            col = 1
            for item in subHeader:
                ws.cell(row=9,column=col,value=item).font = styleBold
                ws.cell(row=9,column=col,value=item).border = thin_border
                ws.cell(row=9,column=col).alignment = Alignment(horizontal='center')
                col = col + 1

            rw = 10
            for item in status:
                fontscheme = ""
                col = 1
                
                for val in item:
                    if col==1:
                        fontscheme = styleBold
                    elif col == 2:
                        fontscheme = styleRight
                    elif col == 3:
                        fontscheme = styleWrong
                    else:
                        fontscheme = styleLight
                    ws.cell(row=rw,column=col,value=val).font = fontscheme
                    ws.cell(row=rw,column=col,value=val).border = thin_border
                    ws.cell(row=rw,column=col).alignment = Alignment(horizontal='center')
                    col = col+1
                rw = rw + 1
            ws.cell(row=12,column=5).font = styleBlue
            ansHeader = ["Student Ans","Correct Ans"]

            rw = 15
            
            ws.cell(row=rw,column=1,value="Student Ans").font=styleBold
            ws.cell(row=rw,column=2,value="Correct Ans").font=styleBold
            ws.cell(row=rw,column=4,value="Student Ans").font=styleBold
            ws.cell(row=rw,column=5,value="Correct Ans").font=styleBold
            ws.cell(row=rw,column=1,value="Student Ans").border = thin_border
            ws.cell(row=rw,column=2,value="Correct Ans").border = thin_border
            ws.cell(row=rw,column=4,value="Student Ans").border = thin_border
            ws.cell(row=rw,column=5,value="Correct Ans").border = thin_border
            ws.cell(row=rw,column=1,value="Student Ans").alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=2,value="Correct Ans").alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=4,value="Student Ans").alignment = Alignment(horizontal='center')
            ws.cell(row=rw,column=5,value="Correct Ans").alignment = Alignment(horizontal='center')
            
            response = value["response"]
            rw = 16
            for i in range(0,len(response)-3):
                studAns = response[i][0]
                corrAns = response[i][1]
                colorScheme = ""
                if studAns==corrAns:
                    colorScheme = styleRight
                else:
                    colorScheme = styleWrong
                ws.cell(row=rw,column=1,value=studAns).font = colorScheme
                ws.cell(row=rw,column=2,value=corrAns).font = styleBlue
                ws.cell(row=rw,column=1,value=studAns).alignment = Alignment(horizontal='center')
                ws.cell(row=rw,column=2,value=corrAns).alignment = Alignment(horizontal='center')
                ws.cell(row=rw,column=1,value=studAns).border = thin_border
                ws.cell(row=rw,column=2,value=corrAns).border = thin_border
                rw = rw + 1
            rw = 16
            for i in range(len(response)-3,len(response)):
                studAns = response[i][0]
                corrAns = response[i][1]
                colorScheme = ""
                if studAns==corrAns:
                    colorScheme = styleRight
                else:
                    colorScheme = styleWrong
                ws.cell(row=rw,column=4,value=studAns).font = colorScheme
                ws.cell(row=rw,column=5,value=corrAns).font = styleBlue
                ws.cell(row=rw,column=4,value=studAns).alignment = Alignment(horizontal='center')
                ws.cell(row=rw,column=5,value=corrAns).alignment = Alignment(horizontal='center')
                ws.cell(row=rw,column=4,value=studAns).border = thin_border
                ws.cell(row=rw,column=5,value=corrAns).border = thin_border
                rw = rw + 1

            wb.save(filePath)
    except Exception as e:
        return "error occurred : " + e
                       
studentMail = {}   
def studentMailDetails():
    try:
        j =0
        with open("media/input/responses.csv",'r',newline='') as File:         #the csv file is stored in a File object
                reader=csv.reader(File)                              #csv.reader is used to read a file
                for row in reader:
                    key = row[6]
                    studentMail[key] = [row[1],row[4],row[3],row[6]]
        
    except Exception as e:
        return "error occurred : " + e

def sendemail():
    try:
        #j=0
        studentMailDetails()
        for filename in os.listdir("output"):
            # j = j + 1
            # if j == 2:
            #     break
            key = re.split(".xlsx",filename)[0]
            if not key in studentMail:
                continue
            student = studentMail[key]
            studentEmail = student[0]
            studentIITPEmail = student[1]
            studentName = student[2]
            msg = EmailMessage()
            msg['Subject'] = "Mail"
            msg['To'] = f'{studentEmail},{studentIITPEmail}'
            msg.set_content(f"Hii , \n {studentName} \n\n Please find the attached quiz marksheet \n\n Regards \n\n Abhay Tiwari & Anurag Victor Ratre")
            with open ("output/"+filename, "rb") as f:
                file_data=f.read ()
                file_name=f.name
                msg.add_attachment(file_data, maintype="application", subtype= "xlsx", filename=file_name)
            with smtplib.SMTP_SSL('smtp.gmail.com',465) as server:
                server.login("victorratreanurag@gmail.com","victor1234569876")
                server.send_message(msg)
                server.quit()
    except Exception as e:
        return e